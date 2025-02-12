import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
import collections
import copy
import inspect
from typing import Type, List, Dict, Any, Optional, Union
import datetime


class ExcelUtil:
    # 用于匹配公式开头的字符串
    FORMULA_REGEX_STR = "=|-|\\+|@"
    FORMULA_STR = ["=", "-", "+", "@"]
    # Excel sheet默认最大行数（这里可以根据实际需求调整，示例中设为和Java代码中类似的较大值）
    sheetSize = 65536
    # 定义thin_border变量，用于设置表格边框样式
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    def __init__(self, clazz):
        self.clazz = clazz
        self.sheet_name = None
        self.type = None
        self.workbook = None
        self.sheet = None
        self.styles = {}
        self.data_list = []
        self.fields = []
        self.rownum = 0
        self.title = None
        self.max_height = 0
        self.sub_merged_last_row_num = 0
        self.sub_merged_first_row_num = 1
        self.sub_method = None
        self.sub_fields = []
        self.statistics = collections.defaultdict(float)
        self.include_fields = []
        self.exclude_fields = []

    def show_column(self, *fields):
        """
        仅在Excel中显示列属性
        """
        self.include_fields = list(fields)

    def hide_column(self, *fields):
        """
        隐藏Excel中列属性
        """
        self.exclude_fields = list(fields)

    def init(self, data_list, sheet_name, title, type_):
        """
        初始化相关属性，为后续操作做准备
        """
        if data_list is None:
            data_list = []
        self.data_list = data_list
        self.sheet_name = sheet_name
        self.type = type_
        self.title = title
        self.create_excel_field()
        self.create_workbook()
        self.create_title()
        self.create_sub_head()

    def create_title(self):
        """
        创建excel第一行标题
        """
        if self.title:
            title_row = self.sheet[0]
            title_cell = title_row[0]
            title_cell.value = self.title
            title_cell.style = self.styles.get("title")
            self.sheet.merge_cells(start_row=0, start_column=0, end_row=0, end_column=len(self.fields) - 1)
            if self.is_sub_list():
                self.sheet.merge_cells(start_row=0, start_column=len(self.fields),
                                       end_row=0, end_column=len(self.fields) + len(self.sub_fields) - 1)

    def create_sub_head(self):
        """
        创建对象的子列表名称
        """
        if self.is_sub_list():
            sub_row = self.sheet[1]
            column = 0
            for field_info in self.fields:
                field = field_info[0]
                excel_anno = field_info[1]
                if isinstance(self.data_list[0][field.name], list):
                    cell = sub_row[column]
                    cell.value = excel_anno.name
                    cell.style = self.styles.get(self._format_header_style_key(excel_anno))
                    if len(self.sub_fields) > 1:
                        self.sheet.merge_cells(start_row=1, start_column=column,
                                               end_row=1, end_column=column + len(self.sub_fields) - 1)
                    column += len(self.sub_fields)
                else:
                    cell = sub_row[column]
                    cell.value = excel_anno.name
                    cell.style = self.styles.get(self._format_header_style_key(excel_anno))
                    column += 1

    def import_excel(self, file_path, title_num=0):
        """
        从Excel文件导入数据到列表，模拟原Java代码中从输入流导入的功能
        """
        try:
            df = pd.read_excel(file_path, engine='openpyxl', header=title_num)
            return self._convert_df_to_list(df)
        except Exception as e:
            print(f"导入Excel异常: {e}")
            raise

    def _convert_df_to_list(self, df):
        """
        将读取的DataFrame数据转换为对象列表
        """
        result_list = []
        columns = df.columns.tolist()
        for index, row in df.iterrows():
            entity = self.clazz()
            for field_info in self.fields:
                field = field_info[0]
                excel_anno = field_info[1]
                col_name = excel_anno.name
                if col_name in columns:
                    val = row[col_name]
                    self._set_entity_value(entity, field, excel_anno, val)
            result_list.append(entity)
        return result_list

    def _set_entity_value(self, entity, field, excel_anno, val):
        """
        根据字段类型和注解设置实体对象的属性值
        """
        field_type = field.type
        if field_type == str:
            val = str(val)
            if val.endswith('.0'):
                val = val[:-2]
            elif excel_anno.date_format:
                val = self.parse_date_to_str(excel_anno.date_format, val)
            elif excel_anno.read_converter_exp:
                val = self.reverse_by_exp(str(val), excel_anno.read_converter_exp, excel_anno.separator)
            else:
                val = str(val)
        elif field_type == int:
            val = int(val) if isinstance(val, (int, float)) and val.is_integer() else val
        elif field_type == float:
            val = float(val)
        elif field_type == bool:
            val = bool(val) if isinstance(val, (int, bool)) else val
        elif field_type == datetime.date:
            val = self.parse_date(val) if isinstance(val, (str, float)) else val
        elif field_type == datetime.datetime:
            val = self.parse_date_time(val) if isinstance(val, (str, float)) else val
        elif hasattr(field_type, '__origin__') and field_type.__origin__ == list:
            sub_list = []
            sub_elem_type = field_type.__args__[0]
            if isinstance(val, list):
                for sub_val in val:
                    sub_entity = sub_elem_type()
                    self._set_sub_entity_value(sub_entity, sub_elem_type, sub_val)
                    sub_list.append(sub_entity)
            setattr(entity, field.name, sub_list)
        else:
            val = val
        setattr(entity, field.name, val)

    def _set_sub_entity_value(self, sub_entity, sub_elem_type, sub_val):
        """
        设置子实体对象的属性值（针对包含子列表的情况）
        """
        for sub_field in self.get_sub_fields(sub_elem_type):
            sub_excel_anno = sub_field.annotation
            sub_col_name = sub_excel_anno.name
            sub_value = sub_val[sub_col_name] if isinstance(sub_val, dict) else None
            self._set_entity_value(sub_entity, sub_field, sub_excel_anno, sub_value)

    def export_excel(self, file_path, data_list=None, sheet_name=None, title=None):
        """
        将数据列表导出到Excel文件
        """
        if data_list is None:
            data_list = self.data_list
        if sheet_name is None:
            sheet_name = self.sheet_name
        if title is None:
            title = self.title
        self.init(data_list, sheet_name, title, "EXPORT")
        self.write_sheet()
        self.workbook.save(file_path)

    def write_sheet(self):
        """
        创建写入数据到Sheet
        """
        sheet_no = max(1, len(self.data_list) // self.sheetSize + 1)
        for index in range(sheet_no):
            self.create_sheet(sheet_no, index)
            header_row = self.sheet[0]
            column = 0
            for field_info in self.fields:
                field = field_info[0]
                excel_anno = field_info[1]
                if isinstance(self.data_list[0][field.name], list):
                    for sub_field in self.sub_fields:
                        sub_excel_anno = sub_field.annotation
                        self.create_head_cell(sub_excel_anno, header_row, column)
                        column += 1
                else:
                    self.create_head_cell(excel_anno, header_row, column)
                    column += 1
            if self.type == "EXPORT":
                self.fill_excel_data(index, header_row)
                self.add_statistics_row()

    def fill_excel_data(self, index, header_row):
        """
        填充excel数据
        """
        start_no = index * self.sheetSize
        end_no = min(start_no + self.sheetSize, len(self.data_list))
        current_row_num = 1  # 从标题行后开始
        for i in range(start_no, end_no):
            row = self.sheet[current_row_num]
            entity = self.data_list[i]
            column = 0
            max_sub_list_size = self.get_current_max_sub_list_size(entity)
            for field_info in self.fields:
                field = field_info[0]
                excel_anno = field_info[1]
                if isinstance(self.data_list[0][field.name], list):
                    sub_list = getattr(entity, field.name)
                    if sub_list:
                        sub_index = 0
                        for sub_entity in sub_list:
                            sub_row = self.sheet[current_row_num + sub_index]
                            if sub_row is None:
                                sub_row = self.sheet[current_row_num + sub_index] = [None] * len(self.sub_fields)
                            sub_column = column
                            for sub_field in self.sub_fields:
                                sub_excel_anno = sub_field.annotation
                                self.add_cell(sub_excel_anno, sub_row, sub_entity, sub_field, sub_column)
                                sub_column += 1
                            sub_index += 1
                    column += len(self.sub_fields)
                else:
                    self.add_cell(excel_anno, row, entity, field, column)
                    if max_sub_list_size > 1 and excel_anno.need_merge:
                        self.sheet.merge_cells(start_row=current_row_num, start_column=column,
                                               end_row=current_row_num + max_sub_list_size - 1, end_column=column)
                    column += 1
            current_row_num += max_sub_list_size

    def get_current_max_sub_list_size(self, entity):
        """
        获取子列表最大数
        """
        max_sub_list_size = 1
        for field_info in self.fields:
            field = field_info[0]
            if isinstance(self.data_list[0][field.name], list):
                sub_list = getattr(entity, field.name)
                if sub_list:
                    max_sub_list_size = max(max_sub_list_size, len(sub_list))
        return max_sub_list_size

    def create_styles(self):
        """
        创建表格样式
        """
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        # 标题样式
        title_style = openpyxl.styles.NamedStyle(name="title")
        title_font = Font(name="Arial", size=16, bold=True)
        title_style.font = title_font
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        title_style.border = thin_border
        self.styles["title"] = title_style

        # 数据样式
        data_style = openpyxl.styles.NamedStyle(name="data")
        data_font = Font(name="Arial", size=10)
        data_style.font = data_font
        data_style.alignment = Alignment(horizontal='center', vertical='center')
        data_style.border = thin_border
        self.styles["data"] = data_style

        # 合计样式
        total_style = openpyxl.styles.NamedStyle(name="total")
        total_font = Font(name="Arial", size=10)
        total_style.font = total_font
        total_style.alignment = Alignment(horizontal='center', vertical='center')
        self.styles["total"] = total_style

        self.styles.update(self.annotation_header_styles())
        self.styles.update(self.annotation_data_styles())

        return self.styles

    def annotation_header_styles(self):
        """
        根据Excel注解创建表格头样式
        """
        header_styles = {}
        for field_info in self.fields:
            excel_anno = field_info[1]
            key = self._format_header_style_key(excel_anno)
            if key not in header_styles:
                style = copy.deepcopy(self.styles["data"])
                style.fill = PatternFill(fill_type="solid", fgColor=excel_anno.header_background_color)
                style.font = Font(name="Arial", size=10, bold=True, color=excel_anno.header_color)
                style.alignment = Alignment(horizontal='center', vertical='center')
                header_styles[key] = style
        return header_styles

    def annotation_data_styles(self):
        """
        根据Excel注解创建表格列样式
        """
        styles = {}
        for field_info in self.fields:
            field = field_info[0]
            excel_anno = field_info[1]
            if isinstance(self.data_list[0][field.name], list):
                sub_elem_type = field.type.__args__[0]
                sub_fields = self.get_sub_fields(sub_elem_type)
                for sub_field in sub_fields:
                    sub_excel_anno = sub_field.annotation
                    self.annotation_data_styles_for_sub_field(styles, sub_field, sub_excel_anno)
            else:
                self.annotation_data_styles_for_field(styles, field, excel_anno)
        return styles

    def annotation_data_styles_for_sub_field(self, styles, sub_field, sub_excel_anno):
        """
        根据Excel注解创建子表格列样式
        """
        key = self._format_data_style_key(sub_excel_anno)
        if key not in styles:
            style = copy.deepcopy(self.styles["data"])
            style.alignment = Alignment(sub_excel_anno.align, vertical='center')
            style.border = self.thin_border
            style.fill = PatternFill(fill_type="solid", fgColor=sub_excel_anno.background_color)
            style.font = Font(name="Arial", size=10, color=sub_excel_anno.color)
            style.number_format = self._get_number_format(sub_excel_anno.cell_type)
            styles[key] = style

    def annotation_data_styles_for_field(self, styles, field, excel_anno):
        """
        根据Excel注解创建表格列样式（针对普通字段）
        """
        key = self._format_data_style_key(excel_anno)
        if key not in styles:
            style = copy.deepcopy(self.styles["data"])
            style.alignment = Alignment(excel_anno.align, vertical='center')
            style.border = self.thin_border
            style.fill = PatternFill(fill_type="solid", fgColor=excel_anno.background_color)
            style.font = Font(name="Arial", size=10, color=excel_anno.color)
            style.number_format = self._get_number_format(excel_anno.cell_type)
            styles[key] = style

    def _format_header_style_key(self, excel_anno):
        """
        格式化表格头样式键
        """
        return f"header_{excel_anno.header_color}_{excel_anno.header_background_color}"

    def _format_data_style_key(self, excel_anno):
        """
        格式化表格列样式键
        """
        return f"data_{excel_anno.align}_{excel_anno.color}_{excel_anno.background_color}_{excel_anno.cell_type}_{excel_anno.wrap_text}"

    def _get_number_format(self, cell_type):
        """
        根据单元格类型获取数字格式
        """
        if cell_type == "TEXT":
            return "@"
        return None

    def create_head_cell(self, excel_anno, row, column):
        """
        创建表格头单元格
        """
        cell = row[column]
        cell.value = excel_anno.name
        cell.style = self.styles.get(self._format_header_style_key(excel_anno))
        self.set_data_validation(excel_anno, row, column)
        if self.is_sub_list():
            self.sheet.column_dimensions[get_column_letter(column + 1)].width = self._get_column_width(excel_anno)
            if excel_anno.need_merge:
                self.sheet.merge_cells(start_row=0, start_column=column, end_row=0, end_column=column)

    def set_data_validation(self, excel_anno, row, column):
        """
        设置单元格信息（如数据验证、列宽等）
        """
        if "注：" in excel_anno.name:
            self.sheet.column_dimensions[get_column_letter(column + 1)].width = 6000
        else:
            self.sheet.column_dimensions[get_column_letter(column + 1)].width = self._get_column_width(excel_anno)
        if excel_anno.prompt or excel_anno.combo:
            if len(excel_anno.combo) > 15 or len("".join(excel_anno.combo)) > 255:
                self.set_xssf_validation_with_hidden(self.sheet, excel_anno.combo, excel_anno.prompt, 1, 100, column, column)
            else:
                self.set_prompt_or_validation(self.sheet, excel_anno.combo, excel_anno.prompt, 1, 100, column, column)

    def add_cell(self, excel_anno, row, entity, field, column):
        """
        添加单元格并设置值等相关操作
        """
        cell = row[column]
        if excel_anno.is_export:
            value = getattr(entity, field.name)
            date_format = excel_anno.date_format
            read_converter_exp = excel_anno.read_converter_exp
            separator = excel_anno.separator
            if date_format and value:
                cell.value = self.parse_date_to_str(date_format, value)
            elif read_converter_exp and value:
                cell.value = self.convert_by_exp(str(value), read_converter_exp, separator)
            elif isinstance(value, float) and excel_anno.scale!= -1:
                cell.value = round(value, excel_anno.scale)
            else:
                self.set_cell_vo(value, excel_anno, cell)
            self.add_statistics_data(column, str(value), excel_anno)
        cell.style = self.styles.get(self._format_data_style_key(excel_anno))

    def set_prompt_or_validation(self, sheet, textlist, prompt_content, first_row, end_row, first_col, end_col):
        """
        设置 POI XSSFSheet 单元格提示或选择框（类似原Java代码功能）
        """
        validation = DataValidation(type="list", formula1='"' + ",".join(textlist) + '"', allow_blank=True)
        if prompt_content:
            validation.promptTitle = ""
            validation.prompt = prompt_content
            validation.showPrompt = True
        validation.add(sheet.cells[first_row - 1:end_row, first_col])
        sheet.add_data_validation(validation)

    def set_xssf_validation_with_hidden(self, sheet, textlist, prompt_content, first_row, end_row, first_col, end_col):
        """
        设置某些列的值只能输入预制的数据，显示下拉框（兼容超出一定数量的下拉框，类似原Java代码功能）
        """
        hide_sheet_name = f"combo_{first_col}_{end_col}"
        hide_sheet = self.workbook.create_sheet(hide_sheet_name)
        for index, text in enumerate(textlist):
            hide_sheet.cell(row=index + 1, column=1, value=text)
        named_range = self.workbook.defined_names.add(f"{hide_sheet_name}_data")
        named_range.value = f"{hide_sheet_name}!$A$1:$A${len(textlist)}"
        validation = DataValidation(type="list", formula1=f"{hide_sheet_name}_data", allow_blank=True)
        if prompt_content:
            validation.promptTitle = ""
            validation.prompt = prompt_content
            validation.showPrompt = True
        validation.add(sheet.cells[first_row - 1:end_row, first_col])
        sheet.add_data_validation(validation)
        self.workbook[hide_sheet_name].sheet_state = "hidden"

    def convert_by_exp(self, property_value, converter_exp, separator):
        """
        解析导出值（类似原Java代码功能）
        """
        result = ""
        convert_source = converter_exp.split(",")
        for item in convert_source:
            item_array = item.split("=")
            if separator in property_value:
                for value in property_value.split(separator):
                    if item_array[0] == value:
                        result += item_array[1] + separator
                        break
            else:
                if item_array[0] == property_value:
                    return item_array[1]
        return result.rstrip(separator)

    def reverse_by_exp(self, property_value, converter_exp, separator):
        """
        反向解析值（类似原Java代码功能）
        """
        result = ""
        convert_source = converter_exp.split(",")
        for item in convert_source:
            item_array = item.split("=")
            if separator in property_value:
                for value in property_value.split(separator):
                    if item_array[1] == value:
                        result += item_array[0] + separator
                        break
            else:
                if item_array[1] == property_value:
                    return item_array[0]
        return result.rstrip(separator)

    def data_format_handler_adapter(self, value, excel_anno, cell):
        """
        数据处理器（类似原Java代码功能，不过Python中实现方式会有差异）
        """
        try:
            handler_class = excel_anno.handler
            handler_instance = handler_class()
            format_method = getattr(handler_instance, "format", None)
            if format_method and inspect.isfunction(format_method):
                value = format_method(value, excel_anno.args, cell, self.workbook)
        except Exception as e:
            print(f"不能格式化数据 {excel_anno.handler}: {e}")
        return value

    def add_statistics_data(self, index, text, excel_anno):
        """
        合计统计信息（类似原Java代码功能）
        """
        if excel_anno and excel_anno.is_statistics:
            try:
                num = float(text)
                self.statistics[index] += num
            except ValueError:
                pass

    def add_statistics_row(self):
        """
        创建统计行（类似原Java代码功能）
        """
        if self.statistics:
            row = self.sheet[self.sheet.max_row + 1]
            cell = row[0]
            cell.value = "合计"
            cell.style = self.styles["total"]
            for index in self.statistics:
                cell = row[index]
                cell.value = round(self.statistics[index], 2)
                cell.style = self.styles["total"]
            self.statistics.clear()

    def get_target_value(self, entity, field, excel_anno):
        """
        获取bean中的属性值（类似原Java代码功能）
        """
        value = getattr(entity, field.name)
        target_attr = excel_anno.target_attr
        if target_attr:
            attrs = target_attr.split(".")
            for attr in attrs:
                value = getattr(value, attr) if hasattr(value, attr) else None
        return value

    def create_excel_field(self):
        """
        得到所有定义字段（类似原Java代码功能）
        """
        self.fields = []
        all_fields = []
        for base in self.clazz.__mro__[:-1]:
            all_fields.extend([(field, getattr(field, 'excel_anno', None)) for field in base.__dict__.values() if isinstance(field, property)])
        if self.include_fields:
            for field, anno in all_fields:
                if field.name in self.include_fields or (anno and isinstance(anno, list) and any(target_attr in self.include_fields for target_attr in [attr.target_attr for attr in anno])):
                    self.fields.append((field, anno))
        elif self.exclude_fields:
            for field, anno in all_fields:
                if field.name not in self.exclude_fields and (not anno or (isinstance(anno, list) and all(target_attr not in self.exclude_fields for target_attr in [attr.target_attr for attr in anno]))):
                    self.fields.append((field, anno))
        else:
            self.fields = [(field, anno) for field, anno in all_fields if anno]
        self.fields.sort(key=lambda x: x[1].sort if x[1] else 0)
        self.max_height = self.get_row_height()

    def get_row_height(self):
        """
        根据注解获取最大行高（类似原Java代码功能）
        """
        max_height = 0
        for field_info in self.fields:
            excel_anno = field_info[1]
            max_height = max(max_height, excel_anno.height)
        return max_height * 20

    def create_workbook(self):
        """
        创建一个工作簿（类似原Java代码功能）
        """
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = self.sheet_name
        self.styles = self.create_styles()

    def create_sheet(self, sheet_no, index):
        """
        创建工作表（类似原Java代码功能）
        """
        if sheet_no > 1 and index > 0:
            self.sheet = self.workbook.create_sheet(title=f"{self.sheet_name}{index}")
            self.create_title()

    def get_cell_value(self, row, column):
        """
        获取单元格值（类似原Java代码功能）
        """
        cell = row[column]
        if cell is None:
            return None
        if isinstance(cell.value, (int, float)):
            if cell.data_type == "n":
                val = cell.value
                if isinstance(val, float) and val.is_integer():
                    val = int(val)
                return val
            elif cell.data_type == "d":
                return cell.value
            elif cell.data_type == "s":
                return str(cell.value)
            elif cell.data_type == "b":
                return bool(cell.value)
            elif cell.data_type == "e":
                return cell.value
        return cell.value

    def is_row_empty(self, row):
        """
        判断是否是空行（类似原Java代码功能）
        """
        for cell in row:
            if cell.value is not None:
                return False
        return True

    def parse_date_to_str(self, date_format, val):
        """
        格式化不同类型的日期对象（类似原Java代码功能）
        """
        if isinstance(val, datetime.datetime):
            return val.strftime(date_format)
        elif isinstance(val, datetime.date):
            return val.strftime(date_format)
        elif isinstance(val, (int, float)):
            date = datetime.datetime.fromtimestamp(val)
            return date.strftime(date_format)
        return str(val)

    def parse_date(self, val):
        """
        解析日期（辅助函数）
        """
        if isinstance(val, str):
            return datetime.datetime.strptime(val, "%Y-%m-%d").date()
        elif isinstance(val, float):
            return datetime.datetime.fromtimestamp(val).date()
        return val

    def parse_date_time(self, val):
        """
        解析日期时间（辅助函数）
        """
        if isinstance(val, str):
            return datetime.datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
        elif isinstance(val, float):
            return datetime.datetime.fromtimestamp(val)
        return val

    def is_sub_list(self):
        """
        是否有对象的子列表（类似原Java代码功能）
        """
        return bool(self.sub_fields)

    def get_sub_fields(self, sub_elem_type):
        """
        获取子列表对应的字段（辅助函数）
        """
        return [field for field in sub_elem_type.__dict__.values() if isinstance(field, property) and hasattr(field, 'excel_anno')]

    def set_cell_vo(self, value, excel_anno, cell):
        """
        设置单元格值（类似原Java代码功能）
        """
        if excel_anno.cell_type in ["STRING", "TEXT"]:
            cell_value = str(value)
            if any(cell_value.startswith(char) for char in self.FORMULA_STR):
                cell_value = re.sub(self.FORMULA_REGEX_STR, "\t\\g<0>", cell_value, 1)
            if isinstance(value, list) and cell_value == "[]":
                cell_value = ""
            cell.value = cell_value + excel_anno.suffix if cell_value else excel_anno.default_value
        elif excel_anno.cell_type == "NUMERIC":
            if value:
                cell.value = float(value) if isinstance(value, (int, float)) and isinstance(value, float) else int(value)
        elif excel_anno.cell_type == "IMAGE":
            # 这里对于图片插入，需要进一步根据具体的图片处理库来完善实现，比如Pillow等，示例中暂不做完整实现
            pass

if __name__ == '__main__':
    import datetime


    # 假设这里有一个简单的示例类，用于演示Excel数据导入导出时对应的数据结构
    class MyData:
        def __init__(self, name, age, sub_list=None):
            self.name = name
            self.age = age
            self.sub_list = sub_list if sub_list is not None else []

        @property
        def excel_anno(self):
            return [
                # 示例注解，可根据实际需求详细配置属性
                {'name': '姓名', 'cell_type': 'TEXT', 'is_export': True},
                {'name': '年龄', 'cell_type': 'NUMERIC', 'is_export': True}
            ]


    # 用于演示子列表中元素的数据结构对应的类（示例）
    class SubData:
        def __init__(self, sub_name, sub_value):
            self.sub_name = sub_name
            self.sub_value = sub_value

        @property
        def excel_anno(self):
            return [
                {'name': '子项名称', 'cell_type': 'TEXT', 'is_export': True},
                {'name': '子项数值', 'cell_type': 'NUMERIC', 'is_export': True}
            ]

    # 测试数据准备
    sub_data1 = SubData("子项1", 10)
    sub_data2 = SubData("子项2", 20)
    my_data1 = MyData("张三", 25, [sub_data1, sub_data2])
    my_data2 = MyData("李四", 30)
    data_list = [my_data1, my_data2]

    # 实例化ExcelUtil
    excel_util = ExcelUtil(MyData)
    # 设置显示的列（这里只是示例，可按需调整）
    excel_util.show_column("name", "age")

    # 测试导出Excel功能
    excel_file_path = "test_export.xlsx"
    excel_util.export_excel(excel_file_path, data_list, sheet_name="测试工作表", title="示例数据")
    print(f"已成功导出Excel文件到 {excel_file_path}")

    # 测试导入Excel功能
    imported_data_list = excel_util.import_excel(excel_file_path)
    print("导入的Excel数据列表:")
    for data in imported_data_list:
        print(f"姓名: {data.name}, 年龄: {data.age}")
        if data.sub_list:
            print("子列表数据:")
            for sub_data in data.sub_list:
                print(f"  子项名称: {sub_data.sub_name}, 子项数值: {sub_data.sub_value}")